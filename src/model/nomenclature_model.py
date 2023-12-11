import pathlib
from fastapi import HTTPException

from starlette.datastructures import UploadFile
from rq.queue import Queue
from rq.job import Job
from redis import Redis
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from datastep.components import datastep_nomenclature
from dto.nomenclature_mapping_job_dto import NomenclatureMappingUpdateDto, NomenclatureMappingJobDto
from infra.supabase import supabase


def parse_txt(file_object: UploadFile) -> (list[tuple], str):
    # Возращаю массив кортежей, в которых на первом месте номенклатура, а на втором — дополнительные инструкции к промпту
    return [(s.decode("utf-8").strip(), "") for s in file_object.file.readlines()], file_object.filename


def parse_sheet(file_object: UploadFile = None) -> (list[tuple], str):
    wb = load_workbook(file_object.file)
    ws = wb.active
    remapping_values = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[7] == "+":
            nomenclature = row[0]
            additional_instructions = row[8]
            remapping_values.append((nomenclature, additional_instructions))
    return remapping_values, file_object.filename


def create_job(test_case: str, filename: str | None, additional_instructions: str = ""):
    split = test_case.split(":")

    query = split[0]
    narrow_group = ""
    middle_group = ""
    wide_group = ""

    if len(split) > 1:
        narrow_group, middle_group = split[1:]

    redis = Redis()
    queue = Queue(name="nomenclature", connection=redis)
    job = queue.enqueue(datastep_nomenclature.do_mapping, query, additional_instructions, result_ttl=-1)

    job.meta["source"] = filename
    job.meta["correct_wide_group"] = wide_group
    job.meta["correct_middle_group"] = middle_group
    job.meta["correct_narrow_group"] = narrow_group
    job.save_meta()


def process(nomenclature_object: UploadFile):
    if nomenclature_object.filename.endswith(".xlsx"):
        test_cases, filename = parse_sheet(nomenclature_object)
    elif nomenclature_object.filename.endswith(".txt"):
        test_cases, filename = parse_txt(nomenclature_object)
    else:
        raise HTTPException(status_code=400, detail="Расширение файла не поддерживается. Используйте txt или xlsx")

    for test_case, additional_instructions in test_cases:
        create_job(test_case, filename, additional_instructions)


def update_nomenclature_mapping(body: NomenclatureMappingUpdateDto):
    supabase.table("nomenclature_mapping").update({"correctness": body.correctness}).eq("id", body.id).execute()


def get_jobs_from_rq(source: str | None) -> list[NomenclatureMappingJobDto]:
    redis = Redis()
    queue = Queue("nomenclature", connection=redis)

    queued_job_ids = queue.get_job_ids()
    started_job_ids = queue.started_job_registry.get_job_ids()
    finished_job_ids = queue.finished_job_registry.get_job_ids()
    failed_job_ids = queue.failed_job_registry.get_job_ids()

    jobs = Job.fetch_many([*queued_job_ids, *started_job_ids, *finished_job_ids, *failed_job_ids], connection=redis)

    result = []
    wanted_jobs = [j for j in jobs if j.get_meta().get("source", None) == source]

    def get_readable_output(output: list[str] | None) -> str:
        if output is None:
            return "None"
        return "\n".join(output)

    for job in wanted_jobs:
        result.append(NomenclatureMappingJobDto(
            id=job.get_meta().get("mapping_id", None),
            input=job.args[0],
            output=get_readable_output(job.return_value()),
            source=job.get_meta().get("source", None),
            status=job.get_status(),
            wide_group=job.get_meta().get("wide_group", None),
            middle_group=job.get_meta().get("middle_group", None),
            narrow_group=job.get_meta().get("narrow_group", None),
            correct_wide_group=job.get_meta().get("correct_wide_group", None),
            correct_middle_group=job.get_meta().get("correct_middle_group", None),
            correct_narrow_group=job.get_meta().get("correct_narrow_group", None)
        ))

    return result


def get_jobs_from_database(source: str | None) -> list[NomenclatureMappingJobDto]:
    response = supabase.table("nomenclature_mapping").select("*").order("id").eq("source", source).execute()

    result = []
    for job in response.data:
        result.append(NomenclatureMappingJobDto(**job))

    return result


def get_all_jobs(source: str | None) -> list[NomenclatureMappingJobDto]:
    jobs_from_rq = get_jobs_from_rq(source)
    # jobs_from_database = get_jobs_from_database(source)
    return jobs_from_rq
    # return [*jobs_from_rq, *jobs_from_database]


def transform_jobs_lists_to_dict(job_lists: list[list[NomenclatureMappingJobDto]]) -> dict:
    """
    result example:
        {'Блок для ручной кладки ЦСК-100 400х200х200мм\n':
            [
                Job('Блок газобетонный D600 B3,5 F50 600х200х200мм', 'None', '01. Строительные материалы', '01.07. Кирпич, камень, блоки', '01.07.01. Блоки газосиликатные', 'source_3.txt'),
                Job('Блок газобетонный D600 B3,5 F50 600х200х200мм', 'None', '01. Строительные материалы', '01.07. Кирпич, камень, блоки', '01.07.01. Блоки газосиликатные', 'source_4.txt')
            ],
            ...
        }

    """
    result = {j.input: [] for j in job_lists[0]}
    for job_list in job_lists:
        for job in job_list:
            result[job.input].append(job)
    return result


def create_test_excel(job_dict: dict, colored: bool = False):
    def color_cell(ws, row: int, column: int, input: str, output: str):
        if output is not None and input.strip() in output:
            ws.cell(row, column).fill = PatternFill("solid", start_color="00FF00")
        else:
            ws.cell(row, column).fill = PatternFill("solid", start_color="FF0000")

    wb = Workbook()
    ws = wb.active

    ws.append(("Вход", "Выход", "Статус", "Широкая группа", "Средняя группа", "Узкая группа", "Источник"))
    for i in range(1, 8):
        ws.cell(1, i).font = Font(bold=True)

    shift = 2
    for input, rows in job_dict.items():
        ws.append((input, *rows[0].to_row()))
        ws.cell(shift, 2).alignment = Alignment(wrapText=True)
        if colored:
            color_cell(ws, shift, 2, input, rows[0].output)
            color_cell(ws, shift, 4, rows[0].correct_wide_group, rows[0].wide_group)
            color_cell(ws, shift, 5, rows[0].correct_middle_group, rows[0].middle_group)
            color_cell(ws, shift, 6, rows[0].correct_narrow_group, rows[0].narrow_group)
        shift += 1
        for job in rows[1:]:
            ws.append(("", *job.to_row()))
            ws.cell(shift, 2).alignment = Alignment(wrapText=True)
            if colored:
                color_cell(ws, shift, 2, input, job.output)
                color_cell(ws, shift, 4, job.correct_wide_group, job.wide_group)
                color_cell(ws, shift, 5, job.correct_middle_group, job.middle_group)
                color_cell(ws, shift, 6, job.correct_narrow_group, job.narrow_group)
            shift += 1

    filepath = f"{pathlib.Path(__file__).parent.resolve()}/../../data/sheet.xlsx"
    wb.save(filepath)


if __name__ == "__main__":
    remapping = parse_sheet()
    create_job(remapping[0][0], "remapping_111223_1", remapping[0][1])
    # first_test_jobs = get_all_jobs("test_101123_1.txt")
    # print(len(first_test_jobs))
    # second_test_jobs = get_all_jobs("test_descr.txt")
    # create_test_excel(transform_jobs_lists_to_dict([first_test_jobs]))
    # print(transform_jobs_lists_to_dict([first_test_jobs, second_test_jobs]))
    # create_excel(all_jobs)

    # for job in all_jobs:
    #     print(job.status, job.correctness, job.id)
