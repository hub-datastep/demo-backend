import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re

def sanitize_sheet_name(name):
    """
    Санитизирует имя листа Excel (удаляет запрещенные символы).
    """
    return re.sub(r'[:*?/\\[\]]', '', name)[:31]

def save_to_excel(data, file_path):
    """
    Сохраняет данные в Excel файл с простым форматированием.
    
    :param data: Список данных, где каждый элемент - это [Номенклатура, Группа]
    :param file_path: Путь для сохранения Excel файла
    """
    
    # Создание DataFrame для листа "DATA"
    df_data = pd.DataFrame(data, columns=['Номенклатура', 'Группа'])
    
    # Создание Excel писателя
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Запись листа "DATA"
        df_data.to_excel(writer, sheet_name='DATA', index=False)
        
        # Группировка данных по группам
        grouped = {}
        for nomenclature, group in data:
            if group not in grouped:
                grouped[group] = []
            grouped[group].append(nomenclature)
        
        # Запись отдельных листов для каждой группы
        for group_name, items in grouped.items():
            df_group = pd.DataFrame(items, columns=['Номенклатура'])
            # Санитизация имени листа
            safe_sheet_name = sanitize_sheet_name(group_name)
            df_group.to_excel(writer, sheet_name=safe_sheet_name, index=False)
    
    # Загрузка книги для форматирования
    wb = load_workbook(file_path)
    
    # Форматирование листа "DATA"
    ws = wb['DATA']
    format_header(ws)
    auto_adjust_columns(ws)  # Автоматическая настройка ширины колонок
    auto_adjust_rows(ws)  # Автоматическая настройка высоты строк
    
    # Форматирование остальных листов
    for sheet_name in wb.sheetnames:
        if sheet_name == 'DATA':
            continue
        ws = wb[sheet_name]
        format_header(ws)
        auto_adjust_columns(ws)
        auto_adjust_rows(ws)
    
    # Сохранение отформатированной книги
    wb.save(file_path)

def format_header(ws):
    """
    Форматирует заголовки листа: жирный шрифт, размер 16, серый фон.
    """
    header_font = Font(bold=True, size=16)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

def auto_adjust_columns(ws):
    """
    Автоматически подбирает ширину колонок.
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def auto_adjust_rows(ws):
    """
    Автоматически подбирает высоту строк.
    """
    for row in ws.rows:
        max_height = 15  # Минимальная высота
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                if lines > max_height:
                    max_height = lines * 15  # Пример расчета высоты
        ws.row_dimensions[row[0].row].height = max_height

# Пример использования
if __name__ == "__main__":
    # Пример структуры данных
    data = [
        ["Номенклатура 1", "Группа A"],
        ["Номенклатура 2", "Группа B"],
        ["Номенклатура 3", "Группа A"],
        ["Номенклатура 4", "Группа C"],
        # Добавьте больше данных по необходимости
    ]
    
    save_to_excel(data, "output.xlsx")
