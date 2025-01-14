import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import json
import re

def sanitize_sheet_name(name):
    # Удаление запрещенных символов
    return re.sub(r'[:*?/\\[\]]', '', name)[:31]

def save_to_excel(data, file_path):
    """
    Сохраняет данные в Excel файл с указанным форматированием.

    :param data: Список списков или кортежей в формате [Номенклатура, Группа, словарь_свойств]
    :param file_path: Путь для сохранения Excel файла
    """
    
    # Создание DataFrame для листа "DATA"
    data_list = []
    for item in data:
        nomenclature, group, properties = item
        # Более читабельный формат свойств
        properties_str = '; '.join([f"{k}: {v}" for k, v in properties.items()])
        data_list.append({
            'Номенклатура': nomenclature,
            'Группа': group,
            'Свойства': properties_str
        })
    
    df_data = pd.DataFrame(data_list)
    
    # Создание Excel писателя
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Запись листа "DATA"
        df_data.to_excel(writer, sheet_name='DATA', index=False)
        
        # Группировка данных по группам
        grouped = {}
        for item in data:
            nomenclature, group, properties = item
            if group not in grouped:
                grouped[group] = []
            grouped[group].append({
                'Номенклатура': nomenclature,
                'Группа': group,
                **properties
            })
        
        # Запись отдельных листов для каждой группы
        for group_name, items in grouped.items():
            df_group = pd.DataFrame(items)
            # Санитизация имени листа
            safe_sheet_name = sanitize_sheet_name(group_name)
            df_group.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        
    # Загрузка книги для форматирования
    wb = load_workbook(file_path)
    
    # Форматирование листа "DATA"
    ws = wb['DATA']
    format_header(ws)
    auto_adjust_columns(ws, specific_columns=['C'])  # Предполагаем, что 'Свойства' это колонка C
    auto_adjust_rows(ws)
    
    # Форматирование остальных листов
    for sheet_name in wb.sheetnames:
        if sheet_name == 'DATA':
            continue
        ws = wb[sheet_name]
        format_header(ws)
        auto_adjust_columns(ws)
        auto_adjust_rows(ws)
    
    # Сохранение отформатированной книги
    wb.save(file_path)

# Остальные функции остаются без изменений


def format_header(ws):
    """
    Форматирует заголовки листа: жирный шрифт, размер 16, серый фон.
    """
    header_font = Font(bold=True, size=16)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

def auto_adjust_columns(ws, specific_columns=None):
    """
    Автоматически подбирает ширину колонок.
    
    :param specific_columns: Список букв колонок для специального подбора (например, 'C' для "Свойства")
    """
    if specific_columns is None:
        specific_columns = []
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        if column_letter in specific_columns:
            ws.column_dimensions[column_letter].width = adjusted_width * 2  # Увеличиваем ширину для "Свойств"
        else:
            ws.column_dimensions[column_letter].width = adjusted_width

def auto_adjust_rows(ws):
    """
    Автоматически подбирает высоту строк.
    """
    for row in ws.rows:
        max_height = 15  # Минимальная высота
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                if lines > max_height:
                    max_height = lines * 15  # Пример расчета высоты
        ws.row_dimensions[row[0].row].height = max_height

# Пример использования
if __name__ == "__main__":
    # Пример структуры данных
    data = [
        ["Номенклатура 1", "Группа A", {"Цвет": "Красный", "Размер": "M"}],
        ["Номенклатура 2", "Группа B", {"Вес": "2kg", "Материал": "Хлопок"}],
        ["Номенклатура 3", "Группа A", {"Цвет": "Синий", "Размер": "L"}],
        # Добавьте больше данных по необходимости
    ]
    
    save_to_excel(data, "output.xlsx")
